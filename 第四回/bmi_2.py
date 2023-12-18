import tkinter
import tkinter.messagebox
import openpyxl as excel


def standard_weight(height):
    return round((height - 100) * 0.9, 4)


def calculate_standard_weight():
    height_value = float(height.get())
    standard_weight_value = round(standard_weight(height_value), 2)
    result_label.config(text=f"標準体重は{standard_weight_value:.2f}kgです。")  # 保留两桁小数
    # Excelファイルに書き込む
    write_to_excel(height_value, None, None, standard_weight_value)


def bmi():
    height_value = float(height.get())
    weight_value = float(weight.get())
    bmi_value = round(weight_value / (height_value / 100) ** 2, 2)  # 四舍五入到两个小数点
    return bmi_value


def calculate_bmi():
    bmi_value = bmi()
    if bmi_value is not None:
        tkinter.messagebox.showinfo("BMI", f"BMIは{bmi_value}です。")
        # Excelファイルに書き込む
        write_to_excel(height.get(), weight.get(), bmi_value, None)


def write_to_excel(height_value, weight_value, bmi_value, standard_weight_value):
    try:
        book = excel.load_workbook("Bmi.xlsx")
        ws = book.active
    except FileNotFoundError:
        book = excel.Workbook()
        ws = book.active

    # 写入身高
    ws["A5"] = "身高"
    if height_value is not None:
        ws["A6"] = height_value
    # 写入体重
    ws["B5"] = "体重"
    if weight_value is not None:
        ws["B6"] = weight_value
    # 写入BMI
    ws["A9"] = "BMI"
    if bmi_value is not None:
        ws["A10"] = bmi_value
    # 写入標準体重
    ws["B9"] = "標準体重"
    if standard_weight_value is not None:
        ws["B10"] = standard_weight_value

    # 罫線の指定
    from openpyxl.styles.borders import Border, Side

    border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
    )

    # セルに罫線を設定　細い罫線を全体に
    for rows in ws["A5":"B6"]:
        for cell in rows:
            cell.border = border

    for rows in ws["A9":"B10"]:
        for cell in rows:
            cell.border = border

        # 保存文件
    book.save("Bmi.xlsx")


root = tkinter.Tk()
root.title("BMI")
root.geometry("400x250")

# 身長の入力欄
height = tkinter.Entry(width=5)
height.place(x=10, y=10)
height_label = tkinter.Label(text="cm")
height_label.place(x=70, y=10)

# 体重の入力欄
weight = tkinter.Entry(width=5)
weight.place(x=10, y=60)
weight_label = tkinter.Label(text="kg")
weight_label.place(x=70, y=60)

# 結果表示のラベル
result_label = tkinter.Label(text="", bg="white", width=40, height=2)
result_label.place(x=10, y=190)

# BMIの計算
button = tkinter.Button(text="BMI", command=calculate_bmi)
button.place(x=10, y=110)

# 標準体重の計算
button = tkinter.Button(text="標準体重", command=calculate_standard_weight)
button.place(x=10, y=160)

root.mainloop()
