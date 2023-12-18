import openpyxl as excel

# 新規ブックを作成
book = excel.Workbook()
sheet = book.active
sheet["B2"] = "売上一覧表"

# 結合
sheet.merge_cells("B2:E2")
# コピー
sheet["B3"] = sheet["B2"].value
# 3行目結合解除
sheet.merge_cells("B3:E3")
sheet.unmerge_cells("B3:E3")

# ファイルに保存
book.save("2_cell.xlsx")
