import openpyxl
import glob

bname = "D:/Python Excel/第六回/選抜メンバー02.xlsx"
wb = openpyxl.load_workbook(bname)
ws = wb.active
mnum = ws.cell(ws.max_row, 1).value
bfiles = glob.glob("第六回　　/./*.txt")
for tname in bfiles:
    with open(tname, mode="r") as f:
        for line in f:
            mnum += 1
            nline = [mnum, line.rstrip("i\n")]
            ws.append(nline)
wb.save(bname)
