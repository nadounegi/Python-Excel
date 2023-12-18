import openpyxl as excel

book = excel.Workbook()
sheet = book.active
# B5に4月入力
cell = sheet["B5"]
cell.value = "4月"
# C4 に5入力
cell = sheet["C4"]
cell.value = "5"
# Cell を使用して、4行目、4列目(D4)に10入力
sheet.cell(row=4, column=4, value=10)
# B5 左右中央揃え
from openpyxl.styles.alignment import Alignment

cell = sheet["B5"]
cell.alignment = Alignment(horizontal="center", vertical="center")
# B5 の横幅30、高さ30
sheet.column_dimensions["B"].width = 30
sheet.row_dimensions[5].height = 30
# B5 のフォントをMS明朝、サイズ20、太字、文字色赤に設定
from openpyxl.styles import Font

cell.font = Font(name="ＭＳ 明朝", size=20, bold=True, color="FF0000")
# D4に背景色をピンク系で塗りつぶす
from openpyxl.styles import PatternFill

cell = sheet["D4"]
cell.fill = PatternFill(fill_type="solid", fgColor="FFC0CB")
# C4 に罫線、上下左右 double 色は緑に設定]
from openpyxl.styles.borders import Border, Side

cell = sheet["C4"]
cell.border = Border(
    top=Side(style="double", color="00FF00"),
    right=Side(style="double", color="00FF00"),
    left=Side(style="double", color="00FF00"),
    bottom=Side(style="double", color="00FF00"),
)
book.save("ren1.xlsx")
