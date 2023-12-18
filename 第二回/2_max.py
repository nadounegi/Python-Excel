import openpyxl as excel

# ワークブックを開いてシートを取り出す
book = excel.load_workbook("第二回/test100.xlsx")
sheet = book.active

# 最下行と最右列を得る
print("最下行:", str(sheet.max_row))
print("最右列:", str(sheet.max_column))

# １行目からデータ入ってないか調べる
i = 1
while i > 0:
    if sheet.cell(i, 2).value == None:
        break
    i = i + 1
print("データの空白行は" + str(i))
