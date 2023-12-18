import openpyxl as excel
from openpyxl.styles import PatternFill

# 新規ワークブックを作成
book = excel.Workbook()
# アクティブなワークシアクティブを取得
sheet = book.active

# 問題1
# セルの設定
for i in range(5):
    sheet.cell(row=(i + 1), column=7, value=i + 1)
    i = i + 1

# 問題2
for i in range(5):
    if i % 2 == 0:  # 偶数であれば
        sheet.cell(row=(i + 1), column=8, value=i + 1)
        i = i + 1

# 問題3
for i in range(6):
    if i % 2 != 0:  # 奇数であれば
        sheet.cell(row=(i + 1), column=8, value=i + 1)
        i = i + 1

# 問題4
# 青系の塗りつぶし
fill = PatternFill(patternType="solid", fgColor="6495ED")
# 7列目のセルを塗りつぶし
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=7, max_col=7):
    for cell in row:
        if cell.value is not None:
            cell.fill = fill

# 問題5
total = 0
for row in sheet.iter_rows(min_row=1, max_row=6, min_col=8, max_col=8):
    for cell in row:
        if cell.value is not None:
            total += cell.value
# H7セルに合計値を入力
cell = sheet.cell(row=7, column=8)
cell.value = total
book.save("第二回/ren2.xlsx")

# 問題6
for y in range(1, 6):
    for x in range(1, 6):
        cell = sheet.cell(row=y, column=x)
        cell.value = str(y) + "行" + str(x) + "列"

book.save("第二回/ren2.xlsx")
