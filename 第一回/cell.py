import openpyxl as excel

book = excel.Workbook()
sheet = book.active
sheet["A1"] = "遊ぼう"
sheet.cell(row=2, column=1, value="皆さん遊ぼう")

cell = sheet.cell(row=3, column=1)
cell.value = "皆さん遊ぼう"
book.save("cell.xlsx")
