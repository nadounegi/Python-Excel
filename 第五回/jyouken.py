import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

fname = "第五回/5_tikan.xlsx"
wb = openpyxl.load_workbook(fname)
ws = wb.active
fnt = Font(color="00008B", bold=True)
fill_red = PatternFill(start_color="EE1111", end_color="EE1111", fill_type="solid")
nrule = CellIsRule(operator="greaterThanOrEqual", formula=[30], font=fnt, fill=fill_red)
ws.conditional_formatting.add("c4:c10", nrule)
wb.save(fname)
