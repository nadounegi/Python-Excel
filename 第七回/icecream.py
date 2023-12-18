import tkinter
import openpyxl as excel
from datetime import datetime

topsu = 0


def show_selected(event):  # リスト変更時プログラム
    global gazou2, n
    n = lb.curselection()  # 選択項目のインデックスを取得
    print(lb.get(n))  # リストのindexにはいろんな情報が入っているので最初の「0]だけだけ
    frame = "第七回/アイス画像/" + "ice" + str(n[0]) + ".gif"
    gazou2 = tkinter.PhotoImage(file=frame)
    canvas.create_image(205, 100, image=gazou2)
    hyoji(topsu)


def check():  # チェックボックス変更時プログラム
    global topsu
    topsu = 0
    for i in range(4):
        if bvar[i].get() == False:
            top_img[i] = tkinter.PhotoImage(file="")
            canvas.create_image(205, 100, image="")
        else:  # チェックボックス選択されてるとき表示
            top_fname = "第七回/アイス画像/" + "topping" + str(i + 1) + ".gif"
            top_img[i] = tkinter.PhotoImage(file=top_fname)
            canvas.create_image(205, 100, image=top_img[i])
            topsu += 1
    hyoji(topsu)


def hyoji(topsu):  # 値段表示
    global topkei
    topkei = topsu * 50
    tanka = [300, 350, 400, 450, 500]
    lbtan["text"] = tanka[n[0]]
    lbtop["text"] = topkei
    lbgokei["text"] = topkei + int(tanka[n[0]])


def click_btn():  # ボタンクリック時プログラム
    global topsu, topkei
    topsu = 0
    book = excel.load_workbook("第七回/regi.xlsx")
    now = datetime.now()
    fmt = now.strftime("%Y年%m月%d日" "%H時%M分%S秒")
    sheet = book.active

    min_row = 1
    while min_row > 0:
        if sheet.cell(min_row, 2).value == None:
            break
        min_row += 1

    sheet.cell(row=min_row, column=2).value = fmt
    sheet.cell(row=min_row, column=3).value = lb.get(n)
    sheet.cell(row=min_row, column=4).value = topkei
    sheet.cell(row=min_row, column=5).value = lbgokei["text"]
    book.save("第七回/regi.xlsx")


root = tkinter.Tk()
root.title("アイスクリーム売店")
root.resizable(False, False)
canvas = tkinter.Canvas(root, width=700, height=500)
canvas.pack()

gazou2 = tkinter.PhotoImage(file="第七回/アイス画像/ice1.gif")
canvas.create_image(205, 100, image=gazou2)

gazou1 = tkinter.PhotoImage(file="第七回/アイス画像/corn.gif")
canvas.create_image(200, 160, image=gazou1)

top_img = [None] * 4  # チェックボックスの箱だけ宣言

list_value = tkinter.StringVar()
list_value.set(["ストロベリー 300円", "バニラ 350円", "抹茶 400円", "チョコレート 450円", "チーズケーキ 500円"])
lb = tkinter.Listbox(height=5, listvariable=list_value, selectmode="single")
lb.place(x=400, y=10)
lb.bind(
    "<<ListboxSelect>>",
    show_selected,
)
bvar = [None] * 4
cbtn = [None] * 4
ITEM = ["チョコチップ", "マカデミアナッツ", "コーンフレーク", "ジェリービーンズ"]

for i in range(4):
    bvar[i] = tkinter.BooleanVar()
    bvar[i].set(False)
    cbtn[i] = tkinter.Checkbutton(
        text=ITEM[i],
        font=("Times New Roman", 12),
        variable=bvar[i],
        bg="#dfe",
        command=check,
    )
    cbtn[i].place(x=400, y=160 + 40 * i)

# 説明
lbsetu = tkinter.Label(root, text="トッピングは１つ50円です")
lbsetu.place(x=320, y=140)
# 値段表示
lbhyo = tkinter.Label(root, text="選択肢アイス", font=("HGPゴシックE", 12))
lbhyo.place(x=10, y=10)

lbtan = tkinter.Label(root, text="値段", font=("HGPゴシックE", 12))
lbtan.place(x=10, y=50)

lbtop = tkinter.Label(root, text="トッピング計", font=("HGPゴシックE", 12))
lbtop.place(x=10, y=70)

lbgokei = tkinter.Label(root, text="合計金額", font=("HGPゴシックE", 12))
lbgokei.place(x=10, y=90)

lbkingaku = tkinter.Label(root, text="金額", font=("HGPゴシックE", 12))
lbkingaku.place(x=10, y=120)

button = tkinter.Button(
    root, text="入力", font=("Times New Roman", 18), command=click_btn
)
button.place(x=200, y=300)
root.mainloop()
