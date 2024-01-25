import openpyxl as excel
import tkinter
import random

book = excel.Workbook()
sheet = book.active
sheet["b4"] = "結果"


# Initialize the variables
s1, s2, s3 = 0, 0, 0


def click_btn1():
    global r1, s1
    r1 = random.randint(1, 3)
    label1["text"] = r1
    label1.update()
    s1 = 1
    kekka()


def click_btn2():
    global r2, s2
    r2 = random.randint(1, 3)
    label2["text"] = r2
    label2.update()
    s2 = 1
    kekka()


def click_btn3():
    global r3, s3
    r3 = random.randint(1, 3)
    label3["text"] = r3
    label3.update()
    s3 = 1
    kekka()


def click_btn4():
    global s1, s2, s3
    label1["text"] = ""
    label2["text"] = ""
    label3["text"] = ""
    lbkekka["text"] = ""
    s1 = 0
    s2 = 0
    s3 = 0


def kekka():
    global s1, s2, s3, r1, r2, r3
    if s1 == 1 and s2 == 1 and s3 == 1:
        if r1 == r2 and r2 == r3:
            lbkekka["text"] = "あたり!"
        else:
            lbkekka["text"] = "はずれ!"
    global count
    count = 4


def click_btn5():
    global count
    count += 1
    sheet.cell(row=count, column=2, value=r1)
    sheet.cell(row=count, column=3, value=r2)
    sheet.cell(row=count, column=4, value=r3)
    sheet.cell(row=count, column=5, value=lbkekka["text"])

    book.save("第五回/slot.xlsx")
    print("保存完了")


root = tkinter.Tk()
root.title("slot")
root.resizable(False, False)
root.geometry("400x600")

label1 = tkinter.Label(root, text="?", font=("Times New Roman", 20), bg="white")
label1.place(x=50, y=80)

label2 = tkinter.Label(root, text="?", font=("Times New Roman", 20), bg="white")
label2.place(x=150, y=80)

label3 = tkinter.Label(root, text="?", font=("Times New Roman", 20), bg="white")
label3.place(x=250, y=80)

button1 = tkinter.Button(
    root, text="stop", font=("Times New Roman", 20), command=click_btn1, fg="skyblue"
)
button1.place(x=50, y=200)

button2 = tkinter.Button(
    root, text="stop", font=("Times New Roman", 20), command=click_btn2, fg="skyblue"
)
button2.place(x=150, y=200)

button3 = tkinter.Button(
    root, text="stop", font=("Times New Roman", 20), command=click_btn3, fg="skyblue"
)
button3.place(x=250, y=200)

lbkekka = tkinter.Label(root, text="?", font=("Times New Roman", 20), bg="white")
lbkekka.place(x=150, y=280)

btnclear = tkinter.Button(
    root, text="クリア", font=("Times New Roman", 20), command=click_btn4, fg="skyblue"
)
btnclear.place(x=150, y=350)

btnsave = tkinter.Button(
    root, text="保存", font=("Times New Roman", 20), command=click_btn5, fg="skyblue"
)
btnsave.place(x=150, y=450)

root.mainloop()
