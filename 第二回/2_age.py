import openpyxl as excel
import datetime

# 新規ワークブックを作成
wb = excel.Workbook()
# アクティブなワークシアクティブを取得
sheet = wb.active

# 今年を得る　年だけ必要
thisYear = datetime.date.today().year
# 0から79まで繰り返し
for i in range(80):
    # 満年齢
    age = i
    # 生年
    year = thisYear - i
    # セルを取得して値を設定
    ageCell = sheet.cell(i + 1, 1)
    ageCell.value = str(i) + "才"
    yearCell = sheet.cell(i + 1, 2)
    yearCell.value = str(year) + "年"

# ファイルを保存
wb.save("第二回/2_age.xlsx")
