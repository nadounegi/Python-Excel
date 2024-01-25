from asyncio.windows_events import INFINITE
from cProfile import label
from logging import root
import tkinter as tk
import openpyxl as excel
import win32com.client as com
import os
from openpyxl.styles import PatternFill


def show_selected(event):
    global selected_sya

    selection = lb.curselection()
    if selection:
        index = selection[0]
        syas = ["相沢 沙希", "生田 真一", "井川 春子", "木村 かえ", "近藤 春香"]
        selected_sya = syas[index]


def get_selected_sya():
    return selected_sya


def sya_no():
    syas = ["相沢 沙希", "生田 真一", "井川 春子", "木村 かえ", "近藤 春香"]
    for i in range(len(syas)):
        if syas[i] == selected_sya:
            sya_no = 101 + int(i)
            i += 1
            return sya_no

    
def caculate_salary():
    # 時間内給与
    base_slary =int(wh.get()) *int(hw.get())
    # 給与合計
    total_salary = base_slary + int(ow.get())

    return base_slary, total_salary

info = []

def click_btn():
    base_salary,total_salary = caculate_salary()
    sya_number = sya_no()  # 获取sya_no函数的返回值

    info.append({
'       sya_number': sya_number,
        'selected_sya': get_selected_sya(),
        'hw': hw.get(),
        'wd': wd.get(),
        'wh': wh.get(),
        'base_salary': base_salary,
        'oh': oh.get(),
        'ow': ow.get(),
        'total_salary': total_salary
    })

    book = excel.load_workbook("第九回/kyuyo.xlsx")
    sheet_sya = get_selected_sya()
    sya_bango = sya_no()

    if sheet_sya:
        selectedsheet = book[str(sya_bango)]
        default_sheet = book["給与一覧"]

        # 開始行を取得
        min_row = 3
        while selectedsheet.cell(row=min_row, column=2).value != None:
            min_row += 1

        default_min_row = 3
        while default_sheet.cell(row=default_min_row, column=2).value != None:
            default_min_row += 1

        # データを書き込む
        # 社員番号
        selectedsheet.cell(row=min_row, column=2).value = sya_no()
        default_sheet.cell(row=default_min_row, column=2).value = sya_no()
        #社員名
        selectedsheet.cell(row=min_row, column=3).value = get_selected_sya()
        default_sheet.cell(row=default_min_row, column=3).value = get_selected_sya()
        #時給
        selectedsheet.cell(row=min_row, column=4).value = hw.get()
        default_sheet.cell(row=default_min_row, column=4).value = hw.get()
        #勤務日
        selectedsheet.cell(row=min_row, column=5).value = wd.get()
        default_sheet.cell(row=default_min_row, column=5).value = wd.get()
        #勤務時間
        selectedsheet.cell(row=min_row, column=6).value = wh.get()
        default_sheet.cell(row=default_min_row, column=6).value = wh.get()
        #時間内給与
        base_salary, total_salary = caculate_salary()
        selectedsheet.cell(row=min_row, column=7).value = base_salary
        default_sheet.cell(row=default_min_row, column=7).value = base_salary
        #残業時間
        overtime_hours = float(oh.get())
        selectedsheet.cell(row=min_row, column=8).value = overtime_hours
        default_sheet.cell(row=default_min_row, column=8).value = overtime_hours
        #残業時間が3時間を超えたら赤色にする
        if overtime_hours >= 3:
             red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
             selectedsheet.cell(row=min_row, column=8).fill = red_fill
             default_sheet.cell(row=default_min_row, column=8).fill = red_fill
        #残業代
        selectedsheet.cell(row=min_row, column=9).value = ow.get()
        default_sheet.cell(row=default_min_row, column=9).value = ow.get()
        #給与合計
        selectedsheet.cell(row=min_row, column=10).value = total_salary
        default_sheet.cell(row=default_min_row, column=10).value = total_salary

        # Excelファイルを保存
        book.save("第九回/kyuyo.xlsx")

def set_worksheet_layout(worksheet):
    # 设置工作表为横向，xlLandscape 的值为 2
    worksheet.PageSetup.Orientation = 2
    # 将内容适应一页宽
    worksheet.PageSetup.FitToPagesWide = 1
    # 注释掉 FitToPagesTall 或将其设置为 1
    # worksheet.PageSetup.FitToPagesTall = 1

def convert_to_pdf(excel_path, pdf_path):
    excel = com.Dispatch("Excel.Application")
    excel.Visible = False
    workbook = None

    try:
        workbook = excel.Workbooks.Open(excel_path)

        # 遍历所有工作表并应用布局设置
        for sheet in workbook.Sheets:
            set_worksheet_layout(sheet)

        workbook.ExportAsFixedFormat(0, pdf_path)
    except Exception as e:
        print("转换失败:", e)
    finally:
        if workbook is not None:
            workbook.Close()
        excel.Quit()
def on_convert_to_pdf():
    excel_path = "D:\Python Excel\第九回\kyuyo.xlsx"  # Excel 文件路径
    pdf_path = "D:\Python Excel\第九回\kyuyo.pdf"    # PDF 文件将被保存的路径
    convert_to_pdf(excel_path, pdf_path)
    print("转换完成")
            
root = tk.Tk()
root.title("給与計算")
#   画面サイズ
root.geometry("1500x1500")

# 社員リスト
sl_lb = tk.Label(root, text="社員リスト", font=("New Times Roman", 12, "bold"))
sl_lb.place(x=235, y=20)

# 社員リストボックス
sya_list = tk.StringVar()
sya_list.set(["相沢　沙希", "生田　真一", "井川　春子", "木村　かえ", "近藤　春香"])
lb = tk.Listbox(root, height=5, listvariable=sya_list, selectmode="single")
lb.place(x=350, y=20)
lb.bind("<<ListboxSelect>>", show_selected)

# 勤務日入力欄
wd_lb = tk.Label(root, text="勤務日", font=("New Times Roman", 12, "bold"))
wd_lb.place(x=235, y=210)
wd = tk.Entry(width=15)
wd.place(x=330, y=210)

# 時給入力欄
hw_lb = tk.Label(root, text="時給", font=("New Times Roman", 12, "bold"))
hw_lb.place(x=235, y=250)
hw = tk.Entry(width=15)
hw.place(x=330, y=250)

# 勤務時間入力欄
wh_lb = tk.Label(root, text="勤務時間", font=("New Times Roman", 12, "bold"))
wh_lb.place(x=235, y=290)
wh = tk.Entry(width=15)
wh.place(x=330, y=290)

# 残業時間入力欄
oh_lb = tk.Label(root, text="残業時間", font=("New Times Roman", 12, "bold"))
oh_lb.place(x=235, y=330)
oh = tk.Entry(width=15)
oh.place(x=330, y=330)

#残業代入力欄
ow_lb = tk.Label(root, text="残業代", font=("New Times Roman", 12, "bold"))
ow_lb.place(x=235, y=370)
ow = tk.Entry(width=15)
ow.place(x=330, y=370)



# 入力ボタン
btn = tk.Button(root, text="入力", font=("New Times Roman", 12, "bold"), command=click_btn)
# ボタンの位置を指定
btn.place(x=235, y=450)

convert_pdf_btn = tk.Button(root, text="PDFファイルに転換", command=on_convert_to_pdf)
convert_pdf_btn.place(x=450, y=450)  # 调整按钮位置
root.mainloop()
