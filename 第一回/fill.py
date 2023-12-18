import openpyxl as excel

book = excel.load_workbook("style.xlsx")
sheet = book.active
cell = sheet["f2"]
cell.value = "背景色"

# 背景色の指定
from openpyxl.styles import PatternFill

cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")
book.save("style.xlsx")
