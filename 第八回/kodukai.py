import openpyxl as excel
import tkinter
import win32com.client as com
import os


def show_selected(event):
    global selected_type

    selection = lb.curselection()
    if selection:
        index = selection[0]
        types = ["交通費", "飲食費", "その他"]
        selected_type = types[index]


def get_selected_type():
    return selected_type


def click_btn():
    book = excel.load_workbook("第八回/kodukai.xlsx")
    sheet_type = get_selected_type()

    if sheet_type:
        selectedsheet = book[sheet_type]
        default_sheet = book["おこづかい"]

        # 这里使用max_row来获取最后一行
        min_row = 3
        while selectedsheet.cell(row=min_row, column=1).value != None:
            min_row += 1

        default_min_row = 3
        while default_sheet.cell(row=default_min_row, column=1).value != None:
            default_min_row += 1
        # 将数据写入对应的sheet
        selectedsheet.cell(row=min_row, column=1).value = date.get()
        default_sheet.cell(row=default_min_row, column=1).value = date.get()
        selectedsheet.cell(row=min_row, column=2).value = name.get()
        default_sheet.cell(row=default_min_row, column=2).value = name.get()
        selectedsheet.cell(row=min_row, column=3).value = price.get()
        default_sheet.cell(row=default_min_row, column=3).value = price.get()

        # 保存Excel文件
        book.save("第八回/kodukai.xlsx")


root = tkinter.Tk()
root.title("こづかい")
root.geometry("600x500")

elb = tkinter.Label(root, text="種類")
elb.place(x=10, y=10)
# リストボックス
expense_list = tkinter.StringVar()
expense_list.set(["交通費", "飲食費", "その他"])
lb = tkinter.Listbox(root, height=3, listvariable=expense_list, selectmode="single")
lb.place(x=70, y=10)  # Adjusted y-coordinate
lb.bind("<<ListboxSelect>>", show_selected)

# 日付入力欄
date_label = tkinter.Label(text="日付")
date_label.place(x=10, y=100)
date = tkinter.Entry(width=15)
date.place(x=70, y=100)

# 品名入力欄
name_label = tkinter.Label(text="品名")
name_label.place(x=10, y=140)
name = tkinter.Entry(width=15)
name.place(x=70, y=140)

# 金額入力欄
price_label = tkinter.Label(text="金額")
price_label.place(x=10, y=180)
price = tkinter.Entry(width=15)
price.place(x=70, y=180)

# 入力ボタン
nbtn = tkinter.Button(root, text="入力", command=click_btn)
# ボタンの位置を指定
nbtn.place(x=10, y=220)
root.mainloop()
